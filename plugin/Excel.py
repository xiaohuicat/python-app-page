import os
import openpyxl as op
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

# 提取列部分
def get_column_name(cell_position):
    return ''.join(filter(str.isalpha, cell_position))

def get_resize_image_path(image_path, width, height):
    image = Image.open(image_path)
    image.thumbnail((width, height))
    temp_dir = tempfile.mkdtemp()  # 创建临时目录
    name = os.path.basename(image_path)
    temp_file_path = os.path.join(temp_dir, name)  # 创建临时文件路径
    image.save(temp_file_path)

    return temp_file_path


class Transform:
    def __init__(self, pixels=0):
        self.pixels = pixels
        
    def pixels_to_points(self, pixels, dpi=96):
        # 默认情况下，假设显示器分辨率为 96 ppi (像素/英寸)
        # 如果你的显示器分辨率不同，可以传入 dpi 参数进行调整
        inches = pixels / dpi
        points = inches * 72
        return points

    def points_to_cm(self, points):
        # 在 96 dpi 的分辨率下，一个 12 磅的字体大致等于 0.16875 英寸
        inches = points / 72  # 1 磅 = 1/72 英寸
        cm = inches * 2.54
        return cm

    def pixels_to_cm(self, pixels):
        points = self.pixels_to_points(pixels)
        return self.points_to_cm(points)
    
    def cm_to_pixels(self, cm, dpi=96):
        inches = cm / 2.54
        points = inches * 72
        pixels = points * dpi / 72
        return pixels

    @property
    def cm(self):
        return self.pixels_to_cm(self.pixels)
    @cm.setter
    def cm(self, value):
        self.pixels = self.cm_to_pixels(value)

    @property
    def points(self):
        return self.pixels_to_points(self.pixels)
    @points.setter
    def points(self, value):
        self.pixels = value * 96 / 72

def fill(color):
  return PatternFill(start_color=color, end_color=color, fill_type='solid')


class Excel:
  def __init__(self, filePath=None) -> None:
    self.filePath = filePath
    self.work_book = None
    self.work_sheet = None
    self.cell_data = None
    self.cell_max_len_dict = {}
    
  def load(self, filePath=None, data_only=False):
    if not filePath:
      self.work_book = op.load_workbook(self.filePath, data_only=data_only) if self.filePath and os.path.exists(self.filePath) else Workbook()
    else:
      self.filePath = filePath
      self.work_book = op.load_workbook(filePath, data_only=data_only)
    return self
  
  def sheet(self, name:str='', create=False):
    if name == '':
      self.work_sheet = self.work_book.active
    else:
      if create:
        self.work_sheet = self.work_book[name] if name in self.work_book.sheetnames else self.work_book.create_sheet(name)
      else:
        self.work_sheet = self.work_book[name] if name in self.work_book.sheetnames else None
    return self
    
  def cell(self, row:int | str, column:int | None = None):
    try:
      if row and not column:
        self.cell_data = self.work_sheet[row]
      else:
        self.cell_data = self.work_sheet.cell(row, column)
    except: 
      pass
    return self
  
  def cell_value(self, data=None):
    if data:
      self.cell_data.value = data
      return self
    elif isinstance(data, int) and data == 0:
      self.cell_data.value = data
      return self
    else:
      return self.cell_data.value
    
  def cell_align_center(self):
    self.cell_data.alignment = Alignment(horizontal='center', vertical='center')
    return self

  def cell_auto_column(self, length=None, mode:str='loose' or 'tight'):
    cell_position = self.cell_data.coordinate
    column_name = get_column_name(cell_position)
    _len = length if length else len(str(self.cell_data.value))
    if column_name in self.cell_max_len_dict:
      self.cell_max_len_dict[column_name] = max(self.cell_max_len_dict[column_name], _len)
    else:
      self.cell_max_len_dict[column_name] = _len
    loose_width = (self.cell_max_len_dict[column_name]+2)*1.2      # 设置缓冲
    tight_width = self.cell_max_len_dict[column_name]              # 不设置缓冲
    self.work_sheet.column_dimensions[column_name].width =  loose_width if mode == 'loose' else tight_width
    return self
    
  def cell_font_color(self, color:str):
    if color == "red":
      self.cell_data.font = Font(color="ff0000")
    elif color == "green":
      self.cell_data.font = Font(color="0cb336")
    elif color == "yellow":
      self.cell_data.font = Font(color="#FFFF00")
    elif color == "blue":
      self.cell_data.font = Font(color="0000ff")
    else:
      self.cell_data.font = Font("333333")
    return self
  
  def cell_bg_color(self, color:str):
    if color == "red":
      self.cell_data.fill = fill(color="ff0000")
    elif color == "green":
      self.cell_data.fill = fill(color="0cb336")
    elif color == "yellow":
      self.cell_data.fill = fill(color="#FFFF00")
    elif color == "blue":
      self.cell_data.fill = fill(color="0000ff")
    else:
      self.cell_data.fill = fill("333333")
    return self
    
  def cell_add_image(self, image_path:str, size_w_h:list=[128,128], resize_rate=0):
    width = Transform(size_w_h[0])
    height = Transform(size_w_h[1])
    
    # 在指定单元格插入图片
    if resize_rate <= 0:
        xl_image = XLImage(Image.open(image_path))
    else:
        new_image_path = get_resize_image_path(image_path, width.pixels*resize_rate, height.pixels*resize_rate)
        xl_image = XLImage(Image.open(new_image_path))
    xl_image.width = width.pixels
    xl_image.height = height.pixels
    
    self.work_sheet.add_image(xl_image, self.cell_data.coordinate)

    # 设置单元格大小，使其适应图片
    self.work_sheet.row_dimensions[self.cell_data.row].height = height.points
    self.work_sheet.column_dimensions[self.cell_data.column_letter].width = width.points / 6  # 大约 7.5 个字符宽度适合一张图片

    return self

  def save(self, name:str):
    self.work_book.save(f'{name}.xlsx')
